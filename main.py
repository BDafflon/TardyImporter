import random
from datetime import datetime
from elasticsearch import Elasticsearch
from os import listdir
from os.path import isfile, join
import openpyxl

from dataStructure import DataStructure


def readFile(workbook):
    data = DataStructure()
    print(workbook.sheetnames)
    if 'Events' in workbook.sheetnames:
        sheet=workbook['Events']
        print("Event.........")
        for row in sheet.iter_rows(min_row=1, max_row=3, min_col=2, max_col=3):
            s=""
            for cell in row:
                s+=str(cell.value)+";"

            if 'enr. lancé à' in s.split(";"):
                data.start=int(s.split(";")[1])
            if 'Enr. stoppé à' in s.split(";"):
                data.end=float(s.split(";")[1])

    if 'DataInfo' in workbook.sheetnames:
        print("DataInfo.........")
        sheet = workbook['DataInfo']
        for row in sheet.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):
            s=""
            for cell in row:
                s+=str(cell.value)+";"

            if 'File name' in s.split(";"):
                data.fileName=s.split(";")[1]
            if 'Start time' in s.split(";"):
                date_time_str = s.split(";")[1].split(" ")[0]+" "+s.split(";")[2]
                date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
                data.date=date_time_obj
            if 'Number of channels' in s.split(";"):
                data.nbChannel=s.split(";")[1]
            if 'Sample rate' in s.split(";"):
                data.rate=s.split(";")[1]

    for sh in workbook.sheetnames:
        if sh.startswith("Data1"):
            sheet = workbook[sh]
            print(sh+".........")
            head=[]
            unit=[]
            head.append(sheet['A1'].value)
            head.append(sheet['B1'].value)
            head.append(sheet['C1'].value)

            unit.append(sheet['A2'].value)
            unit.append(sheet['B2'].value)
            unit.append(sheet['C3'].value)

            for row in sheet.iter_rows(min_row=3):
                s = ""
                for cell in row:
                    s += str(cell.value) + ";"
                if len(s.split(";"))>3:
                    data.data.append({ 'Sensor':{'name':head[1],'unit':unit[1]}, 'Date':{'name':head[0],'unit':unit[0]},'Data':float(s.split(";")[1])})
                    data.data.append({ 'Sensor':{'name':head[2],'unit':unit[2]}, 'Date':{'name':head[0],'unit':unit[0]},'Data':float(s.split(";")[2])})

    print(data.str())

    return data






def importData(dir, server_url, server_port,exp):
    files = [readFile(openpyxl.load_workbook(join(dir, f))) for f in listdir(dir) if isfile(join(dir, f)) and f.lower().endswith('xlsx')]

    es = Elasticsearch([{'host': server_url, 'port': server_port}])
    for f in files:
        meta={'fileName':f.fileName,'start':f.start,"end":f.end,"date":f.date,"rate":f.rate}
        for d in f.data:
            d["meta"]=meta
            res = es.index(index=exp, body=d)




if __name__=='__main__':
    dir="./data"
    server_url="localhost"
    server_port="9200"
    experiment="POC1"
    importData(dir,server_url,server_port,experiment)